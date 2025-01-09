from bs4 import BeautifulSoup
from enum import Enum
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
import re
import requests
from typing import List


def crawl_clien()-> List[List]:
    result: List[List] = []

    url = 'https://davelee-fun.github.io/trial/board/news.html'
    post_response = requests.get(url = 'https://davelee-fun.github.io/trial/board/news.html')

    client_soup = BeautifulSoup(post_response.content, 'html.parser')
    notice_rows = client_soup.select('div.list_item.notice')
    post_rows = client_soup.select('div.list_item.symph_row')[:10]

    rows = notice_rows + post_rows
    for index, row in enumerate(rows):
        # Retrieve elements from posts
        title = re.sub(r'\s{3,}', '', row.select_one('div.list_title > a.list_subject').get_text())
        # replies = re.sub(r'\s*\n?\s*', '', row.select_one('div.list_title > a.list_reply.reply_symph').get_text()
        #                      if row.select_one('div.list_title > a.list_reply.reply_symph') else '')
        link = re.sub(r'news.html', row.select_one('div.list_title > a.list_subject[href]')['href'], url)
        
        reply_response = requests.get(link)
        reply_soup = BeautifulSoup(reply_response.content, 'html.parser')

        # Exclude cases with `blocked` class using `pseudo-class`
        comment_rows = reply_soup.select('div.comment_row:not(.blocked)')
        # Performance-wise, limiting the number of retrieved replies per post
        if len(comment_rows) > 5:
            comment_rows = comment_rows[:5]
            
        commenters_and_comments =[]
        for comment_row in comment_rows:
            # TODO in case, commenter name is with an image
            nick_name_selector = 'div.comment_info > div.post_contact > span.contact_name > span.nickname'
            nick_name_selector_alt = 'div.comment_info > div.post_contact > span.contact_name > span.nickname > img[title]'
            commenter = (comment_row.select_one(nick_name_selector).get_text().strip('\n')
                         if comment_row.select_one(nick_name_selector).get_text().strip('\n') 
                         else comment_row.select_one(nick_name_selector_alt)['title'])
            comment = re.sub(r'\n*\t*', '', comment_row.select_one('div.comment_content').get_text())
            commenters_and_comments.append((commenter, comment))
        # print(f'{index + 1}: {title} [{replies}]\n{link}\n{commenters_and_comments}')

        # `post` keeps the order of 'No', 'Title', 'Commenter', 'Comments', 'Link'
        post = [index + 1, title, commenters_and_comments, link]
        result.append(post)

    return result


def write_xlsx_for_clien(file_name: str, sheet_name: str, post_list: List):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_board = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    header_font_style = Font(name='Arial', size=12, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    title_cell_font = Font(color='00000000', underline='single')
    wrap_text_on = Alignment(wrap_text=True)
    cell_color = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type='solid')

    # create a .xlsx
    excel_file = openpyxl.Workbook()

    # load and configure the current sheet
    current_sheet = excel_file.active

    if sheet_name != '':
        current_sheet.title = sheet_name

    # add the header line
    current_sheet.append(['No', 'Title', 'Commenter', 'Comments'])

    # applying styles on the header
    for cell in current_sheet[1]:
        cell.font = header_font_style
        cell.alignment = header_alignment
        cell.border = thick_board
    
    # cofigure sizes of columns
    current_sheet.column_dimensions['A'].width = 5
    current_sheet.column_dimensions['B'].width = 75
    current_sheet.column_dimensions['C'].width = 15
    current_sheet.column_dimensions['D'].width = 100
    current_sheet.column_dimensions['E'].width = 100
    
    # add data per line
    for post in post_list:
        if post[2]:
            for comment_index, commeter_and_comment in enumerate(post[2]):
                if comment_index == 0:
                    temp = [post[0], post[1], commeter_and_comment[0], commeter_and_comment[1]]
                else:
                    temp = ['', '', commeter_and_comment[0], commeter_and_comment[1]]
                current_sheet.append(temp)

                if comment_index == 0:
                    current_sheet[current_sheet.max_row][1].hyperlink = post[3]
                    current_sheet[current_sheet.max_row][1].font = title_cell_font
                    current_sheet[current_sheet.max_row][1].fill = cell_color
        else:
            temp = [post[0], post[1], '', '']
            current_sheet.append(temp)
            current_sheet[current_sheet.max_row][1].hyperlink = post[3]
            current_sheet[current_sheet.max_row][1].font = title_cell_font
            current_sheet[current_sheet.max_row][1].fill = cell_color

    for row in current_sheet.iter_cols(min_row=2, max_row=current_sheet.max_row, min_col=1, max_col=current_sheet.max_column):
        for cell in row: 
            cell.border = thin_border
            cell.alignment = wrap_text_on

    # save the file
    excel_file.save(f'{file_name}.xlsx')
    excel_file.close()


def extract_posts_from_clien():
    write_xlsx_for_clien('clien', 'clien', crawl_clien())
    print('hiw')

extract_posts_from_clien()